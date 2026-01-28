import uuid
import asyncio
import logging
import os
import sys
import re
import random
import sqlite3
import io
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse

from app.db import get_conn, init_db, dumps, loads
from app.schemas import (
    JobStatus,
    Party,
    InvoiceDetail,
    InvoiceListItem,
    RequestRescanRequest,
    UpdateInvoiceRequest,
    UploadResponse,
    UploadJobResponse,
)

BASE_DIR = Path(__file__).resolve().parents[1]
STORAGE_DIR = BASE_DIR / "storage"
STORAGE_DIR.mkdir(parents=True, exist_ok=True)

# Ensure repo root is importable (so we can import sibling invoice_extractor/)
REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))


def _load_env_file(env_path: Path) -> None:
    if not env_path.exists():
        return
    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


_load_env_file(REPO_ROOT / "invoice_extractor" / ".env")
API_KEY = os.getenv("GEMINI_API_KEY")

# Import extractor library from sibling folder (after sys.path fix)
from invoice_extractor.lib import INVOICE_FIELDS, extract_from_pdf_bytes  # noqa: E402

app = FastAPI(title="Invoice Backend", version="0.1.0")

logger = logging.getLogger("invoice-backend")
if not logger.handlers:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s: %(message)s")

def _env_int(name: str, default: int) -> int:
    try:
        v = int(str(os.getenv(name, "")).strip() or default)
        return v
    except Exception:
        return default


# Performance tuning (safe defaults)
# - DPI 200 is usually enough for printed invoices and much faster than 300+.
# - Retry renders run only when first pass is low-readability.
# - Concurrency 2 is a good balance on flaky networks (fast + fewer disconnects).
EXTRACT_DPI = max(100, min(_env_int("INVOICE_EXTRACT_DPI", 200), 400))
RETRY_DPI = max(EXTRACT_DPI, min(_env_int("INVOICE_RETRY_DPI", 300), 600))
MAX_PAGE_CONCURRENCY = max(1, min(_env_int("INVOICE_PAGE_CONCURRENCY", 3), 3))
PAGE_TIMEOUT_S = max(30, min(_env_int("INVOICE_PAGE_TIMEOUT_S", 180), 900))
BIG_PDF_PAGES = max(1, min(_env_int("INVOICE_BIG_PDF_PAGES", 10), 500))
BIG_PDF_BYTES = max(1024 * 1024, min(_env_int("INVOICE_BIG_PDF_BYTES", 8 * 1024 * 1024), 200 * 1024 * 1024))
VLM_BATCH_SIZE = max(1, min(_env_int("INVOICE_VLM_BATCH_SIZE", 3), 6))
VLM_IMAGE_FORMAT = (os.getenv("INVOICE_VLM_IMAGE_FORMAT", "jpeg") or "jpeg").strip().lower()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://127.0.0.1:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def _startup():
    init_db()


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


async def _extract_pages_with_retry(
    *,
    pdf_bytes: bytes,
    only_page: int | None = None,
    page_numbers: list[int] | None = None,
    dpi: int = EXTRACT_DPI,
    retry_dpi: int = RETRY_DPI,
):
    """
    Run extraction off the event loop (thread) and retry transient Gemini failures.
    This prevents the whole API server from "freezing" during model calls.
    """
    if not API_KEY:
        raise HTTPException(status_code=500, detail="GEMINI_API_KEY not configured on backend.")

    # Retry/backoff tuned for flaky networks + transient upstream issues.
    # Jitter helps avoid thundering herd when multiple pages retry together.
    backoff_s = [1.0, 2.0, 4.0, 8.0, 16.0]
    last_err: Exception | None = None

    for attempt in range(len(backoff_s) + 1):
        try:
            return await asyncio.wait_for(
                asyncio.to_thread(
                    extract_from_pdf_bytes,
                    pdf_bytes,
                    api_key=API_KEY,
                    dpi=dpi,
                    model_name="gemini-2.5-flash",
                    only_page=only_page,
                    page_numbers=page_numbers,
                    render_retry=True,  # only triggers zoom-crops when first pass is low readability
                    retry_dpi=retry_dpi,
                    batch_size=1 if only_page is not None else VLM_BATCH_SIZE,
                    image_format=VLM_IMAGE_FORMAT,
                ),
                timeout=PAGE_TIMEOUT_S,
            )
        except Exception as e:  # noqa: BLE001 - we normalize upstream errors here
            last_err = e
            msg = str(e) or ""
            msg_u = msg.upper()
            msg_l = msg.lower()

            is_503 = ("503" in msg) or ("UNAVAILABLE" in msg_u)
            is_transient_net = any(
                s in msg_l
                for s in [
                    "getaddrinfo failed",
                    "name or service not known",
                    "temporary failure in name resolution",
                    "errno 11001",
                    "server disconnected",
                    "unexpected_eof_while_reading",
                    "eof occurred in violation of protocol",
                    "connection reset",
                    "connection aborted",
                    "timed out",
                ]
            )
            is_timeout = isinstance(e, asyncio.TimeoutError) or ("timeout" in msg_l)
            is_retryable = is_503 or is_transient_net or is_timeout

            if is_retryable and attempt < len(backoff_s):
                base = backoff_s[attempt]
                delay = base * random.uniform(0.8, 1.3)
                logger.warning(
                    "Extraction retryable error (attempt %s), retrying in %.1fs (page=%s): %s",
                    attempt + 1,
                    delay,
                    only_page,
                    msg,
                )
                await asyncio.sleep(delay)
                continue

            if is_503:
                logger.exception("Gemini unavailable after retries")
                raise HTTPException(
                    status_code=503,
                    detail="Gemini service temporarily unavailable. Please retry in a minute.",
                )

            logger.exception("Extraction failed")
            raise

    # unreachable, but keeps type-checkers happy
    raise last_err  # type: ignore[misc]

def _job_row_to_model(row) -> JobStatus:
    keys = set(row.keys())

    def _opt(col: str):
        return row[col] if col in keys else None

    invoice_ids = loads(_opt("invoice_ids_json")) or []
    if not isinstance(invoice_ids, list):
        invoice_ids = []

    return JobStatus(
        id=row["id"],
        document_id=row["document_id"],
        filename=_opt("filename"),
        status=row["status"],
        total_pages=_opt("total_pages"),
        processed_pages=int(row["processed_pages"] or 0),
        message=_opt("message"),
        error=_opt("error"),
        invoice_ids=invoice_ids,
        has_low_readability=bool(row["has_low_readability"]),
        created_at=_opt("created_at"),
        updated_at=_opt("updated_at"),
    )


def _update_job(
    *,
    job_id: str,
    status: str | None = None,
    total_pages: int | None = None,
    processed_pages: int | None = None,
    message: str | None = None,
    error: str | None = None,
    invoice_ids: list[str] | None = None,
    has_low_readability: bool | None = None,
) -> None:
    fields: list[str] = []
    vals: list[Any] = []

    if status is not None:
        fields.append("status = ?")
        vals.append(status)
    if total_pages is not None:
        fields.append("total_pages = ?")
        vals.append(total_pages)
    if processed_pages is not None:
        fields.append("processed_pages = ?")
        vals.append(processed_pages)
    if message is not None:
        fields.append("message = ?")
        vals.append(message)
    if error is not None:
        fields.append("error = ?")
        vals.append(error)
    if invoice_ids is not None:
        fields.append("invoice_ids_json = ?")
        vals.append(dumps(invoice_ids))
    if has_low_readability is not None:
        fields.append("has_low_readability = ?")
        vals.append(1 if has_low_readability else 0)

    fields.append("updated_at = ?")
    vals.append(_now_iso())

    conn = get_conn()
    try:
        conn.execute(
            f"UPDATE jobs SET {', '.join(fields)} WHERE id = ?",
            (*vals, job_id),
        )
        conn.commit()
    finally:
        conn.close()


def _is_low_readability(p) -> bool:
    """
    True only when the page has actual readability/vision issues,
    not merely "needs review" due to missing/optional fields.
    """
    reasons = set((getattr(p, "reasons", None) or []) if isinstance(getattr(p, "reasons", None), list) else [])
    if any(r.startswith("json_parse_failed") for r in reasons):
        return True
    # Only treat explicit "unreadable" signals as low readability.
    # "low_field_confidence" and "render_retry_used" can happen on perfectly readable pages
    # (model being conservative), so they should not force a Low Readability status.
    if "model_flagged_unreadable" in reasons:
        return True

    diag = getattr(p, "field_diagnostics", None) or {}
    if isinstance(diag, dict):
        critical = {"Invoice_No", "Invoice_Date", "Net_Amount"}
        for f in critical:
            v = diag.get(f)
            if not isinstance(v, dict):
                continue
            status = str(v.get("status") or "").lower()
            if status not in {"unreadable", "blurry", "faded", "cut_off"}:
                continue
            # Date "format ambiguity" is not a readability problem per user requirement.
            if f == "Invoice_Date":
                reason = str(v.get("reason") or "").lower()
                if any(s in reason for s in ["format", "invalid day", "invalid month", "ambiguous date"]):
                    continue
            return True

    # If the model asked for rescan AND it couldn't confidently read critical fields, mark low readability.
    unreadable_fields = set(getattr(p, "unreadable_fields", None) or [])
    if bool(getattr(p, "needs_rescan", False)) and (unreadable_fields & {"Invoice_No", "Invoice_Date", "Net_Amount"}):
        return True
    return False


def _insert_invoice_for_page(conn, *, doc_id: str, p) -> tuple[str, bool, bool]:
    """
    Insert a single invoice row for a page extraction.
    Returns (invoice_id, needs_review, low_readability)
    """
    inv_id = uuid.uuid4().hex
    extracted = dict(p.data)
    now = _now_iso()

    system_conf = p.system_confidence if p.system_confidence is not None else 1.0
    diag = p.field_diagnostics or {}
    needs_audit = any(
        isinstance(v, dict) and v.get("status") in {"ambiguous"}
        for v in diag.values()
    )
    low_readability = _is_low_readability(p)
    needs_review = bool(p.needs_rescan) or bool(needs_audit) or (system_conf < 0.85)
    status = "needs-review" if needs_review else "auto-extracted"

    supplier_party_id = _upsert_party(
        conn,
        party_type="supplier",
        name=extracted.get("Supplier_Name"),
        ntn=extracted.get("Supplier_NTN"),
        gst_no=extracted.get("Supplier_GST_No"),
        registration_no=extracted.get("Supplier_Registration_No"),
    )
    buyer_party_id = _upsert_party(
        conn,
        party_type="buyer",
        name=extracted.get("Buyer_Name"),
        ntn=extracted.get("Buyer_NTN"),
        gst_no=extracted.get("Buyer_GST_No"),
        registration_no=extracted.get("Buyer_Registration_No"),
    )

    conn.execute(
        """
        INSERT INTO invoices
          (id, document_id, page_no, supplier_party_id, buyer_party_id, extracted_json, edited_json, status, needs_rescan,
           unreadable_fields_json, reasons_json, model_avg_confidence, system_confidence,
           system_reasons_json, field_diagnostics_json, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            inv_id,
            doc_id,
            int(getattr(p, "page_no", 0) or 0),
            supplier_party_id,
            buyer_party_id,
            dumps(extracted),
            None,
            status,
            1 if p.needs_rescan else 0,
            dumps(p.unreadable_fields),
            dumps(p.reasons),
            getattr(p, "avg_field_confidence", None),
            getattr(p, "system_confidence", None),
            dumps(getattr(p, "system_reasons", []) or []),
            dumps(getattr(p, "field_diagnostics", {}) or {}),
            now,
            now,
        ),
    )

    return inv_id, needs_review, low_readability


async def _run_document_job(*, job_id: str, doc_id: str, stored_path: Path) -> None:
    """
    Background job that extracts pages and writes invoices incrementally.
    """
    try:
        _update_job(job_id=job_id, status="running", message="Reading PDF...")
        pdf_bytes = stored_path.read_bytes()
        import fitz  # local import to keep startup light

        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        total_pages = len(doc)
        doc.close()

        _update_job(job_id=job_id, total_pages=total_pages, processed_pages=0, message="Extracting pages...")

        invoice_ids: list[str] = []
        any_low_readability = False
        processed_pages = 0

        # Preflight: extract page 1 first to assess readability; if it's big AND low-readability,
        # switch to concurrency=1 to reduce disconnects/rate-limit pressure.
        _update_job(job_id=job_id, message=f"Extracting pages... (0/{total_pages})")

        first_pages = await _extract_pages_with_retry(
            pdf_bytes=pdf_bytes,
            only_page=1,
            dpi=EXTRACT_DPI,
            retry_dpi=RETRY_DPI,
        )
        first = first_pages[0] if first_pages else None
        if first is None:
            raise RuntimeError("Extraction returned no pages for page 1")

        conn = get_conn()
        try:
            inv_id, _needs_review, low_readability = _insert_invoice_for_page(conn, doc_id=doc_id, p=first)
            conn.commit()
        finally:
            conn.close()

        invoice_ids.append(inv_id)
        any_low_readability = any_low_readability or low_readability
        processed_pages = 1
        _update_job(
            job_id=job_id,
            processed_pages=processed_pages,
            invoice_ids=invoice_ids,
            has_low_readability=any_low_readability,
            message=f"Extracting pages... ({processed_pages}/{total_pages})",
        )

        effective_concurrency = MAX_PAGE_CONCURRENCY
        if (total_pages >= BIG_PDF_PAGES or len(pdf_bytes) >= BIG_PDF_BYTES) and any_low_readability:
            effective_concurrency = 1
            _update_job(
                job_id=job_id,
                message=f"Low-readability big PDF detected; switching to sequential processing (1/{total_pages})",
            )

        sem = asyncio.Semaphore(effective_concurrency)
        db_lock = asyncio.Lock()  # serialize sqlite writes to avoid "database is locked"
        state_lock = asyncio.Lock()

        def _chunks(seq: list[int], n: int) -> list[list[int]]:
            n = max(1, int(n or 1))
            return [seq[i : i + n] for i in range(0, len(seq), n)]

        async def _process_batch(page_nums_1based: list[int]) -> None:
            nonlocal any_low_readability, processed_pages
            if not page_nums_1based:
                return
            async with sem:
                pages = await _extract_pages_with_retry(
                    pdf_bytes=pdf_bytes,
                    page_numbers=page_nums_1based,
                    dpi=EXTRACT_DPI,
                    retry_dpi=RETRY_DPI,
                )
                if not pages:
                    raise RuntimeError(f"Extraction returned no pages for batch {page_nums_1based}")

                # Write invoices to DB (serialized).
                inv_ids: list[str] = []
                batch_low = False
                async with db_lock:
                    conn = get_conn()
                    try:
                        for p in pages:
                            inv_id, _needs_review, low_readability = _insert_invoice_for_page(conn, doc_id=doc_id, p=p)
                            inv_ids.append(inv_id)
                            batch_low = batch_low or low_readability
                        conn.commit()
                    finally:
                        conn.close()

                # Update shared job state/progress.
                async with state_lock:
                    invoice_ids.extend(inv_ids)
                    any_low_readability = any_low_readability or batch_low
                    processed_pages += len(inv_ids)
                    _update_job(
                        job_id=job_id,
                        processed_pages=processed_pages,
                        invoice_ids=invoice_ids,
                        has_low_readability=any_low_readability,
                        message=f"Extracting pages... ({processed_pages}/{total_pages})",
                    )

        remaining = list(range(2, total_pages + 1))
        batches = _chunks(remaining, VLM_BATCH_SIZE)
        tasks = [asyncio.create_task(_process_batch(b)) for b in batches]
        try:
            await asyncio.gather(*tasks)
        except Exception:
            for t in tasks:
                t.cancel()
            raise

        _update_job(
            job_id=job_id,
            status="completed",
            message="Completed",
            processed_pages=total_pages,
            invoice_ids=invoice_ids,
            has_low_readability=any_low_readability,
        )
        logger.info("Job completed job_id=%s document_id=%s pages=%s", job_id, doc_id, total_pages)
    except Exception as e:  # noqa: BLE001
        logger.exception("Job failed job_id=%s document_id=%s", job_id, doc_id)
        _update_job(job_id=job_id, status="failed", error=str(e), message="Failed")


@app.get("/api/jobs/{job_id}", response_model=JobStatus)
def get_job(job_id: str):
    conn = get_conn()
    try:
        row = conn.execute(
            """
            SELECT j.*, d.filename AS filename
              FROM jobs j
              JOIN documents d ON d.id = j.document_id
             WHERE j.id = ?
            """,
            (job_id,),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Job not found.")
        return _job_row_to_model(row)
    finally:
        conn.close()


@app.get("/api/jobs", response_model=list[JobStatus])
def list_jobs(limit: int = 50):
    limit = max(1, min(int(limit or 50), 200))
    conn = get_conn()
    try:
        rows = conn.execute(
            """
            SELECT j.*, d.filename AS filename
              FROM jobs j
              JOIN documents d ON d.id = j.document_id
             ORDER BY j.created_at DESC
             LIMIT ?
            """,
            (limit,),
        ).fetchall()
        return [_job_row_to_model(r) for r in rows]
    finally:
        conn.close()

_ID_PLACEHOLDER_RAW = {
    "N/A",
    "NA",
    "N.A",
    "N.A.",
    "NOT APPLICABLE",
    "NOTAPPLICABLE",
    "NONE",
    "NULL",
}


def _norm_id(value: Any) -> str:
    """
    Normalize government identifiers (NTN/GST/Registration) for matching.
    Keep alphanumerics only, uppercase.
    """
    if value is None:
        return ""
    raw = str(value).strip()
    if not raw:
        return ""
    upper = raw.upper()
    if upper in _ID_PLACEHOLDER_RAW:
        return ""
    s = upper
    s = "".join(ch for ch in s if ch.isalnum())
    return s


def _norm_name(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip().lower()
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"[^a-z0-9 ]", "", s)
    return s.strip()


def _upsert_party(
    conn,
    *,
    party_type: str,
    name: Any,
    ntn: Any,
    gst_no: Any,
    registration_no: Any,
) -> str:
    """
    Match/merge rule:
    - If NTN exists -> match on (type, ntn_norm)
    - else if Registration exists -> match on (type, registration_norm)
    - else -> create a new party (name-only identities are not reliable)
    """
    now = _now_iso()

    name_raw = (str(name).strip() if name is not None else None)
    name_norm = _norm_name(name) or None

    ntn_raw = (str(ntn).strip() if ntn is not None else None)
    ntn_norm = _norm_id(ntn) or None

    gst_raw = (str(gst_no).strip() if gst_no is not None else None)
    gst_norm = _norm_id(gst_no) or None

    reg_raw = (str(registration_no).strip() if registration_no is not None else None)
    reg_norm = _norm_id(registration_no) or None

    def _find_by(col: str, norm_val: str | None) -> str | None:
        if not norm_val:
            return None
        row = conn.execute(
            f"SELECT id FROM parties WHERE type = ? AND {col} = ? LIMIT 1",
            (party_type, norm_val),
        ).fetchone()
        return row["id"] if row else None

    party_id = (
        _find_by("ntn_norm", ntn_norm)
        or _find_by("registration_norm", reg_norm)
    )

    # If both identifiers exist but point to different rows, merge them deterministically.
    id_by_ntn = _find_by("ntn_norm", ntn_norm)
    id_by_reg = _find_by("registration_norm", reg_norm)
    if id_by_ntn and id_by_reg and id_by_ntn != id_by_reg:
        # Prefer NTN as the canonical key (more stable/less ambiguous).
        canonical_id = id_by_ntn
        other_id = id_by_reg
        # Move registration identifier onto canonical row by clearing it from the other row first.
        conn.execute(
            """
            UPDATE parties
              SET registration_raw = NULL,
                  registration_norm = NULL,
                  updated_at = ?
            WHERE id = ?
            """,
            (now, other_id),
        )
        conn.execute(
            """
            UPDATE parties
              SET name_raw = COALESCE(?, name_raw),
                  name_norm = COALESCE(?, name_norm),
                  ntn_raw = COALESCE(?, ntn_raw),
                  ntn_norm = COALESCE(?, ntn_norm),
                  gst_raw = COALESCE(?, gst_raw),
                  gst_norm = COALESCE(?, gst_norm),
                  registration_raw = COALESCE(?, registration_raw),
                  registration_norm = COALESCE(?, registration_norm),
                  updated_at = ?
            WHERE id = ?
            """,
            (
                name_raw,
                name_norm,
                ntn_raw,
                ntn_norm,
                gst_raw,
                gst_norm,
                reg_raw,
                reg_norm,
                now,
                canonical_id,
            ),
        )
        return canonical_id

    if party_id:
        try:
            conn.execute(
                """
                UPDATE parties
                  SET name_raw = COALESCE(?, name_raw),
                      name_norm = COALESCE(?, name_norm),
                      ntn_raw = COALESCE(?, ntn_raw),
                      ntn_norm = COALESCE(?, ntn_norm),
                      gst_raw = COALESCE(?, gst_raw),
                      gst_norm = COALESCE(?, gst_norm),
                      registration_raw = COALESCE(?, registration_raw),
                      registration_norm = COALESCE(?, registration_norm),
                      updated_at = ?
                WHERE id = ?
                """,
                (
                    name_raw,
                    name_norm,
                    ntn_raw,
                    ntn_norm,
                    gst_raw,
                    gst_norm,
                    reg_raw,
                    reg_norm,
                    now,
                    party_id,
                ),
            )
            return party_id
        except sqlite3.IntegrityError:
            # This happens when we matched an existing party, but applying *new* identifiers
            # (e.g. registration_norm) would collide with another party row.
            # Prefer the canonical row for that identifier.
            canonical_id = _find_by("ntn_norm", ntn_norm) or _find_by("registration_norm", reg_norm)
            if canonical_id and canonical_id != party_id:
                # Ensure uniqueness by "moving" identifiers to canonical row.
                if reg_norm:
                    owner = _find_by("registration_norm", reg_norm)
                    if owner and owner != canonical_id:
                        conn.execute(
                            """
                            UPDATE parties
                              SET registration_raw = NULL,
                                  registration_norm = NULL,
                                  updated_at = ?
                            WHERE id = ?
                            """,
                            (now, owner),
                        )
                if ntn_norm:
                    owner = _find_by("ntn_norm", ntn_norm)
                    if owner and owner != canonical_id:
                        conn.execute(
                            """
                            UPDATE parties
                              SET ntn_raw = NULL,
                                  ntn_norm = NULL,
                                  updated_at = ?
                            WHERE id = ?
                            """,
                            (now, owner),
                        )
                try:
                    conn.execute(
                        """
                        UPDATE parties
                          SET name_raw = COALESCE(?, name_raw),
                              name_norm = COALESCE(?, name_norm),
                              ntn_raw = COALESCE(?, ntn_raw),
                              ntn_norm = COALESCE(?, ntn_norm),
                              gst_raw = COALESCE(?, gst_raw),
                              gst_norm = COALESCE(?, gst_norm),
                              registration_raw = COALESCE(?, registration_raw),
                              registration_norm = COALESCE(?, registration_norm),
                              updated_at = ?
                        WHERE id = ?
                        """,
                        (
                            name_raw,
                            name_norm,
                            ntn_raw,
                            ntn_norm,
                            gst_raw,
                            gst_norm,
                            reg_raw,
                            reg_norm,
                            now,
                            canonical_id,
                        ),
                    )
                except sqlite3.IntegrityError:
                    # As a last resort, don't persist conflicting identifiers—keep the canonical row.
                    conn.execute(
                        """
                        UPDATE parties
                          SET name_raw = COALESCE(?, name_raw),
                              name_norm = COALESCE(?, name_norm),
                              gst_raw = COALESCE(?, gst_raw),
                              gst_norm = COALESCE(?, gst_norm),
                              updated_at = ?
                        WHERE id = ?
                        """,
                        (name_raw, name_norm, gst_raw, gst_norm, now, canonical_id),
                    )
                return canonical_id
            raise

    party_id = uuid.uuid4().hex
    try:
        conn.execute(
            """
            INSERT INTO parties
              (id, type, name_raw, name_norm, ntn_raw, ntn_norm, gst_raw, gst_norm, registration_raw, registration_norm, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (party_id, party_type, name_raw, name_norm, ntn_raw, ntn_norm, gst_raw, gst_norm, reg_raw, reg_norm, now, now),
        )
        return party_id
    except sqlite3.IntegrityError:
        # Another concurrent request/job likely inserted the same party identifier.
        # Re-select the existing row and update it with any new non-null details.
        existing_id = _find_by("ntn_norm", ntn_norm) or _find_by("registration_norm", reg_norm)
        if not existing_id:
            raise
        try:
            conn.execute(
                """
                UPDATE parties
                  SET name_raw = COALESCE(?, name_raw),
                      name_norm = COALESCE(?, name_norm),
                      ntn_raw = COALESCE(?, ntn_raw),
                      ntn_norm = COALESCE(?, ntn_norm),
                      gst_raw = COALESCE(?, gst_raw),
                      gst_norm = COALESCE(?, gst_norm),
                      registration_raw = COALESCE(?, registration_raw),
                      registration_norm = COALESCE(?, registration_norm),
                      updated_at = ?
                WHERE id = ?
                """,
                (
                    name_raw,
                    name_norm,
                    ntn_raw,
                    ntn_norm,
                    gst_raw,
                    gst_norm,
                    reg_raw,
                    reg_norm,
                    now,
                    existing_id,
                ),
            )
        except sqlite3.IntegrityError:
            # If the update would attach an identifier owned by a different row, "move" it.
            if reg_norm:
                owner = _find_by("registration_norm", reg_norm)
                if owner and owner != existing_id:
                    conn.execute(
                        """
                        UPDATE parties
                          SET registration_raw = NULL,
                              registration_norm = NULL,
                              updated_at = ?
                        WHERE id = ?
                        """,
                        (now, owner),
                    )
            if ntn_norm:
                owner = _find_by("ntn_norm", ntn_norm)
                if owner and owner != existing_id:
                    conn.execute(
                        """
                        UPDATE parties
                          SET ntn_raw = NULL,
                              ntn_norm = NULL,
                              updated_at = ?
                        WHERE id = ?
                        """,
                        (now, owner),
                    )
            conn.execute(
                """
                UPDATE parties
                  SET name_raw = COALESCE(?, name_raw),
                      name_norm = COALESCE(?, name_norm),
                      gst_raw = COALESCE(?, gst_raw),
                      gst_norm = COALESCE(?, gst_norm),
                      registration_raw = COALESCE(?, registration_raw),
                      registration_norm = COALESCE(?, registration_norm),
                      ntn_raw = COALESCE(?, ntn_raw),
                      ntn_norm = COALESCE(?, ntn_norm),
                      updated_at = ?
                WHERE id = ?
                """,
                (
                    name_raw,
                    name_norm,
                    gst_raw,
                    gst_norm,
                    reg_raw,
                    reg_norm,
                    ntn_raw,
                    ntn_norm,
                    now,
                    existing_id,
                ),
            )
        return existing_id


def _invoice_list_row_to_model(row) -> InvoiceListItem:
    keys = set(row.keys())
    def _opt(col: str):
        return row[col] if col in keys else None

    extracted = loads(row["extracted_json"]) or {}
    edited = loads(row["edited_json"]) or {}
    current = edited or extracted
    return InvoiceListItem(
        id=row["id"],
        document_id=row["document_id"],
        page_no=int(row["page_no"]),
        supplier_party_id=_opt("supplier_party_id"),
        buyer_party_id=_opt("buyer_party_id"),
        status=row["status"],
        needs_rescan=bool(row["needs_rescan"]),
        unreadable_fields=loads(row["unreadable_fields_json"]) or [],
        reasons=loads(row["reasons_json"]) or [],
        extracted=extracted,
        current=current,
        model_avg_confidence=_opt("model_avg_confidence"),
        system_confidence=_opt("system_confidence"),
        system_reasons=loads(_opt("system_reasons_json")) or [],
        field_diagnostics=loads(_opt("field_diagnostics_json")) or {},
    )


@app.post("/api/documents", response_model=UploadJobResponse)
async def upload_document(file: UploadFile = File(...)):
    if file.content_type not in ("application/pdf", "application/x-pdf", "application/acrobat", "applications/vnd.pdf"):
        # browsers often send application/pdf; keep relaxed but require .pdf name too
        if not (file.filename or "").lower().endswith(".pdf"):
            raise HTTPException(status_code=400, detail="Only PDF uploads are supported.")

    if not API_KEY:
        raise HTTPException(status_code=500, detail="GEMINI_API_KEY not configured on backend.")

    doc_id = uuid.uuid4().hex
    stored_path = STORAGE_DIR / f"{doc_id}.pdf"

    pdf_bytes = await file.read()
    stored_path.write_bytes(pdf_bytes)

    created_at = _now_iso()
    conn = get_conn()
    try:
        conn.execute(
            "INSERT INTO documents (id, filename, stored_path, created_at) VALUES (?, ?, ?, ?)",
            (doc_id, file.filename or "upload.pdf", str(stored_path), created_at),
        )
        conn.commit()
    finally:
        conn.close()

    # Create a background job and return immediately for progress polling.
    job_id = uuid.uuid4().hex
    now = _now_iso()
    conn = get_conn()
    try:
        conn.execute(
            """
            INSERT INTO jobs (id, document_id, status, total_pages, processed_pages, message, error, invoice_ids_json, has_low_readability, created_at, updated_at)
            VALUES (?, ?, 'queued', NULL, 0, ?, NULL, ?, 0, ?, ?)
            """,
            (job_id, doc_id, "Queued", dumps([]), now, now),
        )
        conn.commit()
    finally:
        conn.close()

    logger.info("Enqueued job job_id=%s document_id=%s filename=%s", job_id, doc_id, file.filename)
    asyncio.create_task(_run_document_job(job_id=job_id, doc_id=doc_id, stored_path=stored_path))

    return UploadJobResponse(job_id=job_id, document_id=doc_id)


@app.get("/api/documents/{document_id}/file")
def get_document_file(document_id: str):
    conn = get_conn()
    try:
        row = conn.execute("SELECT stored_path, filename FROM documents WHERE id = ?", (document_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Document not found.")
        path = row["stored_path"]
        name = row["filename"]
    finally:
        conn.close()

    # IMPORTANT: open inline in browser (do not force download)
    headers = {"Content-Disposition": f'inline; filename="{name}"'}
    return FileResponse(path, media_type="application/pdf", headers=headers)


@app.get("/api/invoices", response_model=list[InvoiceListItem])
def list_invoices(include_history: bool = False):
    conn = get_conn()
    try:
        if include_history:
            rows = conn.execute("SELECT * FROM invoices ORDER BY created_at DESC").fetchall()
        else:
            # Default: show only the most recent document per filename.
            # This avoids duplicates when the same PDF is uploaded multiple times.
            rows = conn.execute(
                """
                WITH latest_docs AS (
                  SELECT d1.id
                    FROM documents d1
                    JOIN (
                      SELECT filename, MAX(created_at) AS max_created_at
                        FROM documents
                       GROUP BY filename
                    ) m
                      ON m.filename = d1.filename
                     AND m.max_created_at = d1.created_at
                )
                SELECT i.*
                  FROM invoices i
                 WHERE i.document_id IN (SELECT id FROM latest_docs)
                 ORDER BY i.created_at DESC
                """
            ).fetchall()
        return [_invoice_list_row_to_model(r) for r in rows]
    finally:
        conn.close()


@app.get("/api/invoices/export.xlsx")
def export_invoices_xlsx(include_history: bool = False):
    """
    Backend-generated Excel export of invoices.
    By default, matches the invoice list behavior (latest document per filename).
    """
    try:
        from openpyxl import Workbook
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Excel export dependency missing (openpyxl). {e}")

    conn = get_conn()
    try:
        if include_history:
            rows = conn.execute(
                """
                SELECT i.*, d.filename
                  FROM invoices i
                  JOIN documents d ON d.id = i.document_id
                 ORDER BY i.created_at DESC
                """
            ).fetchall()
        else:
            rows = conn.execute(
                """
                WITH latest_docs AS (
                  SELECT d1.id
                    FROM documents d1
                    JOIN (
                      SELECT filename, MAX(created_at) AS max_created_at
                        FROM documents
                       GROUP BY filename
                    ) m
                      ON m.filename = d1.filename
                     AND m.max_created_at = d1.created_at
                )
                SELECT i.*, d.filename
                  FROM invoices i
                  JOIN documents d ON d.id = i.document_id
                 WHERE i.document_id IN (SELECT id FROM latest_docs)
                 ORDER BY i.created_at DESC
                """
            ).fetchall()
    finally:
        conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    # Export should match the user's CSV template exactly (columns + order).
    # Source: `Jalal Sons (1).csv`
    headers = [
        "Invoice_Date",
        "Invoice_No",
        "Supplier_Name",
        "Supplier_NTN",
        "Supplier_GST_No",
        "Supplier_Registration_No",
        "Buyer_Name",
        "Buyer_NTN",
        "Buyer_GST_No",
        "Buyer_Registration_No",
        "Exclusive_Value",
        "GST_Sales_Tax",
        "Inclusive_Value",
        "Advance_Tax",
        "Net_Amount",
        "Return",
        "Discount",
        "Incentive",
        "Location",
        "GRN",
    ]
    ws.append(headers)

    for r in rows:
        extracted = loads(r["extracted_json"]) or {}
        edited = loads(r["edited_json"]) or {}
        current = edited or extracted
        ws.append(
            [
                current.get("Invoice_Date"),
                current.get("Invoice_No"),
                current.get("Supplier_Name"),
                current.get("Supplier_NTN"),
                current.get("Supplier_GST_No"),
                current.get("Supplier_Registration_No"),
                current.get("Buyer_Name"),
                current.get("Buyer_NTN"),
                current.get("Buyer_GST_No"),
                current.get("Buyer_Registration_No"),
                current.get("Exclusive_Value"),
                current.get("GST_Sales_Tax"),
                current.get("Inclusive_Value"),
                current.get("Advance_Tax"),
                current.get("Net_Amount"),
                current.get("Return"),
                current.get("Discount"),
                current.get("Incentive"),
                current.get("Location"),
                current.get("GRN"),
            ]
        )

    # Make it readable in Excel: freeze header row + light autosizing.
    ws.freeze_panes = "A2"
    try:
        for col_idx, col_name in enumerate(headers, start=1):
            max_len = len(col_name)
            for row_idx in range(2, min(ws.max_row, 250) + 1):  # cap scan for speed
                v = ws.cell(row=row_idx, column=col_idx).value
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 45)
    except Exception:
        # sizing is best-effort
        pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"invoices_{ts}.xlsx"
    headers_out = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_out,
    )


@app.get("/api/ai-review", response_model=list[InvoiceListItem])
def list_ai_review(include_history: bool = False):
    conn = get_conn()
    try:
        if include_history:
            rows = conn.execute(
                "SELECT * FROM invoices WHERE needs_rescan = 1 OR status = 'needs-review' ORDER BY updated_at DESC"
            ).fetchall()
        else:
            rows = conn.execute(
                """
                WITH latest_docs AS (
                  SELECT d1.id
                    FROM documents d1
                    JOIN (
                      SELECT filename, MAX(created_at) AS max_created_at
                        FROM documents
                       GROUP BY filename
                    ) m
                      ON m.filename = d1.filename
                     AND m.max_created_at = d1.created_at
                )
                SELECT i.*
                  FROM invoices i
                 WHERE (i.needs_rescan = 1 OR i.status = 'needs-review')
                   AND i.document_id IN (SELECT id FROM latest_docs)
                 ORDER BY i.updated_at DESC
                """
            ).fetchall()
        return [_invoice_list_row_to_model(r) for r in rows]
    finally:
        conn.close()


@app.get("/api/invoices/{invoice_id}", response_model=InvoiceDetail)
def get_invoice(invoice_id: str):
    conn = get_conn()
    try:
        row = conn.execute("SELECT * FROM invoices WHERE id = ?", (invoice_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Invoice not found.")
        keys = set(row.keys())
        def _opt(col: str):
            return row[col] if col in keys else None
        extracted = loads(row["extracted_json"]) or {}
        edited = loads(row["edited_json"]) or {}
        current = edited or extracted
        document_id = row["document_id"]
        page_no = int(row["page_no"])
    finally:
        conn.close()

    # Use PDF viewer page jump (works in most browsers): .../file#page=2
    document_url = f"/api/documents/{document_id}/file#page={page_no}"

    return InvoiceDetail(
        id=row["id"],
        document_id=document_id,
        page_no=page_no,
        supplier_party_id=_opt("supplier_party_id"),
        buyer_party_id=_opt("buyer_party_id"),
        status=row["status"],
        needs_rescan=bool(row["needs_rescan"]),
        unreadable_fields=loads(row["unreadable_fields_json"]) or [],
        reasons=loads(row["reasons_json"]) or [],
        extracted=extracted,
        edited=edited or extracted,
        current=current,
        model_avg_confidence=_opt("model_avg_confidence"),
        system_confidence=_opt("system_confidence"),
        system_reasons=loads(_opt("system_reasons_json")) or [],
        field_diagnostics=loads(_opt("field_diagnostics_json")) or {},
        document_url=document_url,
    )


@app.get("/api/parties", response_model=list[Party])
def list_parties(party_type: str | None = None):
    conn = get_conn()
    try:
        if party_type:
            rows = conn.execute(
                "SELECT * FROM parties WHERE type = ? ORDER BY name_norm, id",
                (party_type,),
            ).fetchall()
        else:
            rows = conn.execute("SELECT * FROM parties ORDER BY type, name_norm, id").fetchall()

        out: list[Party] = []
        for r in rows:
            out.append(
                Party(
                    id=r["id"],
                    type=r["type"],
                    name=r["name_raw"],
                    ntn=r["ntn_raw"],
                    gst_no=r["gst_raw"],
                    registration_no=r["registration_raw"],
                    created_at=r["created_at"],
                    updated_at=r["updated_at"],
                )
            )
        return out
    finally:
        conn.close()


@app.get("/api/parties/{party_id}", response_model=Party)
def get_party(party_id: str):
    conn = get_conn()
    try:
        r = conn.execute("SELECT * FROM parties WHERE id = ?", (party_id,)).fetchone()
        if not r:
            raise HTTPException(status_code=404, detail="Party not found.")
        return Party(
            id=r["id"],
            type=r["type"],
            name=r["name_raw"],
            ntn=r["ntn_raw"],
            gst_no=r["gst_raw"],
            registration_no=r["registration_raw"],
            created_at=r["created_at"],
            updated_at=r["updated_at"],
        )
    finally:
        conn.close()


@app.put("/api/invoices/{invoice_id}", response_model=InvoiceDetail)
def update_invoice(invoice_id: str, payload: UpdateInvoiceRequest):
    conn = get_conn()
    try:
        row = conn.execute("SELECT * FROM invoices WHERE id = ?", (invoice_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Invoice not found.")

        extracted = loads(row["extracted_json"]) or {}
        edited = loads(row["edited_json"]) or extracted

        if payload.edited is not None and isinstance(payload.edited, dict):
            # keep only known fields
            next_edited: dict[str, Any] = {k: payload.edited.get(k) for k in INVOICE_FIELDS}
            edited = next_edited

        status = payload.status or row["status"]
        now = _now_iso()

        # approving clears rescan flag
        needs_rescan = row["needs_rescan"]
        if status == "approved":
            needs_rescan = 0

        conn.execute(
            "UPDATE invoices SET edited_json = ?, status = ?, needs_rescan = ?, updated_at = ? WHERE id = ?",
            (dumps(edited), status, needs_rescan, now, invoice_id),
        )
        conn.commit()
    finally:
        conn.close()

    # return updated detail
    return get_invoice(invoice_id)


@app.post("/api/invoices/{invoice_id}/request-rescan", response_model=InvoiceDetail)
def request_rescan(invoice_id: str, payload: RequestRescanRequest):
    conn = get_conn()
    try:
        row = conn.execute("SELECT * FROM invoices WHERE id = ?", (invoice_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Invoice not found.")

        existing_unreadable = loads(row["unreadable_fields_json"]) or []
        existing_reasons = loads(row["reasons_json"]) or []

        unreadable = sorted(set(existing_unreadable + (payload.unreadable_fields or [])))
        reasons = sorted(set(existing_reasons + (payload.reasons or []) + ["user_requested_rescan"]))

        now = _now_iso()
        conn.execute(
            """
            UPDATE invoices
              SET needs_rescan = 1,
                  status = 'needs-review',
                  unreadable_fields_json = ?,
                  reasons_json = ?,
                  updated_at = ?
            WHERE id = ?
            """,
            (dumps(unreadable), dumps(reasons), now, invoice_id),
        )
        conn.commit()
    finally:
        conn.close()

    return get_invoice(invoice_id)


@app.post("/api/invoices/{invoice_id}/reupload", response_model=InvoiceDetail)
async def reupload_invoice(invoice_id: str, file: UploadFile = File(...)):
    """
    Replace the document backing an invoice and re-run extraction.
    Keeps the same invoice_id (record), but stores a new document and points the invoice to it.
    """
    if not API_KEY:
        raise HTTPException(status_code=500, detail="GEMINI_API_KEY not configured on backend.")

    if file.content_type not in ("application/pdf", "application/x-pdf", "application/acrobat", "applications/vnd.pdf"):
        if not (file.filename or "").lower().endswith(".pdf"):
            raise HTTPException(status_code=400, detail="Only PDF uploads are supported.")

    conn = get_conn()
    try:
        row = conn.execute("SELECT * FROM invoices WHERE id = ?", (invoice_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Invoice not found.")
    finally:
        conn.close()

    pdf_bytes = await file.read()

    # Store new document (do not affect other invoices/pages)
    new_doc_id = uuid.uuid4().hex
    stored_path = STORAGE_DIR / f"{new_doc_id}.pdf"
    stored_path.write_bytes(pdf_bytes)

    now = _now_iso()
    conn = get_conn()
    try:
        conn.execute(
            "INSERT INTO documents (id, filename, stored_path, created_at) VALUES (?, ?, ?, ?)",
            (new_doc_id, file.filename or "reupload.pdf", str(stored_path), now),
        )
        conn.commit()
    finally:
        conn.close()

    # Re-extract only the first page of the reuploaded PDF and use it as the replacement.
    logger.info("Starting reupload extraction for invoice_id=%s new_doc_id=%s bytes=%s", invoice_id, new_doc_id, len(pdf_bytes))
    pages = await _extract_pages_with_retry(pdf_bytes=pdf_bytes, only_page=1)
    logger.info("Finished reupload extraction for invoice_id=%s new_doc_id=%s pages=%s", invoice_id, new_doc_id, len(pages))
    if not pages:
        raise HTTPException(status_code=400, detail="Uploaded PDF has no pages.")
    p = pages[0]

    extracted = dict(p.data)
    status = "needs-review" if p.needs_rescan else "auto-extracted"

    conn = get_conn()
    try:
        conn.execute(
            """
            UPDATE invoices
              SET document_id = ?,
                  page_no = ?,
                  extracted_json = ?,
                  edited_json = NULL,
                  status = ?,
                  needs_rescan = ?,
                  unreadable_fields_json = ?,
                  reasons_json = ?,
                  updated_at = ?
            WHERE id = ?
            """,
            (
                new_doc_id,
                1,
                dumps(extracted),
                status,
                1 if p.needs_rescan else 0,
                dumps(p.unreadable_fields),
                dumps(sorted(set((p.reasons or []) + ["reuploaded_and_reprocessed"]))),
                now,
                invoice_id,
            ),
        )
        conn.commit()
    finally:
        conn.close()

    return get_invoice(invoice_id)

